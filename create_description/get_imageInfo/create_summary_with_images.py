import pandas as pd
import os
import time
import openpyxl
from PIL import Image
import io
import base64
from typing import Dict, List
from azure.identity import InteractiveBrowserCredential, get_bearer_token_provider
from openai import AzureOpenAI

AZURE_OPENAI_ENDPOINT = 'https://aoai-je-exm.openai.azure.com/'
DEPLOYMENT_NAME = 'gpt-4o'  # gpt-4oはvision対応

# 段階1: 各シートの概要を作成
SHEET_SUMMARY_PROMPT = """
以下は半導体露光装置における{basename}という機能仕様書の「{sheet_name}」シートの内容です。
このシートの内容を簡潔に要約してください（200-300文字程度）。

要約には以下を含めてください：
- このシートの目的・役割
- 主要な情報の種類（パラメータ、フロー、状態遷移など）
- 重要なポイント
- 画像が含まれている場合は、その内容も要約に含めてください
"""

# 段階2: 全体の解説を作成
FINAL_SUMMARY_PROMPT = """
以下は半導体露光装置における{basename}という機能仕様書の全シート概要です。
これらの情報を統合して、この機能の包括的な解説を作成してください。

## 解説に含めるべき内容
1. **機能概要**: この機能の目的と役割
2. **主要な処理フロー**: どのような処理が行われるか
3. **入出力**: 何を受け取り、何を出力するか
4. **重要なパラメータ**: キーとなる設定値や制約
5. **関連する機能**: 他の機能との関係性
6. **特記事項**: 注意すべき点や制約条件

回答はMarkdown形式で構造化してください。
"""

INPUT_DIR = r'resource'
OUTPUT_DIR = r'summary'

# レート制限対策の待機時間（秒）
WAIT_TIME_BETWEEN_SHEETS = 2
WAIT_TIME_BETWEEN_FILES = 5

# シートを分割する閾値（行数）
MAX_ROWS_PER_CHUNK = 100


def get_excel_files(dir_path: str) -> list[str]:
    """
    指定ディレクトリ配下のすべてのExcelファイルを再帰的に取得
    ルートディレクトリからの相対パス（拡張子なし）のリストを返す
    """
    excel_files = []
    
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
        return excel_files
    
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, dir_path)
                rel_path_without_ext = os.path.splitext(rel_path)[0]
                excel_files.append(rel_path_without_ext)
    
    return excel_files


def get_client() -> AzureOpenAI:
    credential = InteractiveBrowserCredential(
        tenant_id="4876a51c-4f2d-4d54-b712-e0b67d308e80"
    )
    
    token_provider = get_bearer_token_provider(
        credential,
        "https://cognitiveservices.azure.com/.default"
    )

    return AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        azure_ad_token_provider=token_provider,
        api_version="2024-12-01-preview",
    )


def extract_images_from_sheet(excel_path: str, sheet_name: str) -> List[Image.Image]:
    """
    特定のシートから画像を抽出
    
    Args:
        excel_path: Excelファイルのパス
        sheet_name: シート名
    
    Returns:
        PIL Imageのリスト
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]
        
        images = []
        if hasattr(sheet, '_images'):
            for img in sheet._images:
                image_data = img._data()
                pil_image = Image.open(io.BytesIO(image_data))
                images.append(pil_image)
        
        return images
    except Exception as e:
        print(f'    Warning: Could not extract images from {sheet_name}: {str(e)}')
        return []


def image_to_base64(image: Image.Image) -> str:
    """
    PIL ImageをBase64文字列に変換
    """
    buffered = io.BytesIO()
    # 大きな画像はリサイズ
    max_size = (1024, 1024)
    image.thumbnail(max_size, Image.Resampling.LANCZOS)
    image.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return img_str


def create_message_content(text: str, images: List[Image.Image]) -> list:
    """
    テキストと画像を含むメッセージコンテンツを作成
    """
    content = [{"type": "text", "text": text}]
    
    # 画像を追加（最大5枚まで）
    for img in images[:5]:
        base64_image = image_to_base64(img)
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/png;base64,{base64_image}"
            }
        })
    
    return content


def split_dataframe_into_chunks(df: pd.DataFrame, max_rows: int = MAX_ROWS_PER_CHUNK) -> List[tuple]:
    """DataFrameを指定行数で分割"""
    chunks = []
    total_rows = len(df)
    
    for start in range(0, total_rows, max_rows):
        end = min(start + max_rows, total_rows)
        chunk = df.iloc[start:end]
        chunks.append((start, end, chunk))
    
    return chunks


def summarize_sheet(basename: str, sheet_name: str, df: pd.DataFrame, 
                   excel_path: str, client: AzureOpenAI,
                   use_chunking: bool = False, include_images: bool = True) -> str:
    """
    1つのシートの概要を作成（画像対応版）
    """
    # 画像を抽出
    images = []
    if include_images:
        images = extract_images_from_sheet(excel_path, sheet_name)
        if images:
            print(f'    Found {len(images)} images in sheet')
    
    # チャンク分割が不要な場合
    if not use_chunking or len(df) <= MAX_ROWS_PER_CHUNK:
        table_text = df.to_markdown()
        user_content = create_message_content(table_text, images)
        
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": SHEET_SUMMARY_PROMPT.format(
                    basename=basename, 
                    sheet_name=sheet_name
                )},
                {"role": "user", "content": user_content}
            ],
            max_tokens=1000
        )
        return response.choices[0].message.content
    
    # チャンク分割して処理
    chunks = split_dataframe_into_chunks(df)
    chunk_summaries = []
    
    print(f'    Splitting into {len(chunks)} chunks...')
    
    for i, (start, end, chunk_df) in enumerate(chunks, 1):
        chunk_info = f"行{start+1}-{end}"
        print(f'    Processing chunk {i}/{len(chunks)}: {chunk_info}')
        
        try:
            table_text = chunk_df.to_markdown()
            
            # 最初のチャンクにのみ画像を含める
            chunk_images = images if i == 1 else []
            user_content = create_message_content(table_text, chunk_images)
            
            response = client.chat.completions.create(
                model=DEPLOYMENT_NAME,
                messages=[
                    {"role": "system", "content": SHEET_SUMMARY_PROMPT.format(
                        basename=basename, 
                        sheet_name=f"{sheet_name} ({chunk_info})"
                    )},
                    {"role": "user", "content": user_content}
                ],
                max_tokens=1000
            )
            chunk_summaries.append(response.choices[0].message.content)
            
            if i < len(chunks):
                time.sleep(WAIT_TIME_BETWEEN_SHEETS)
                
        except Exception as e:
            print(f'    ⚠ Error processing chunk {chunk_info}: {str(e)}')
            chunk_summaries.append(f"[{chunk_info}] エラー: {str(e)}")
    
    return "\n\n".join(chunk_summaries)


def create_sheet_summaries(basename: str, all_sheets: Dict[str, pd.DataFrame], 
                          excel_path: str, client: AzureOpenAI, 
                          use_chunking: bool = False, include_images: bool = True) -> Dict[str, str]:
    """
    全シートの概要を作成
    """
    sheet_summaries = {}
    
    for i, (sheet_name, df) in enumerate(all_sheets.items(), 1):
        print(f'  Summarizing sheet {i}/{len(all_sheets)}: {sheet_name} ({len(df)} rows)')
        
        try:
            summary = summarize_sheet(basename, sheet_name, df, excel_path, client, 
                                     use_chunking, include_images)
            sheet_summaries[sheet_name] = summary
            
            if i < len(all_sheets):
                time.sleep(WAIT_TIME_BETWEEN_SHEETS)
                
        except Exception as e:
            print(f'  ⚠ Error summarizing sheet {sheet_name}: {str(e)}')
            sheet_summaries[sheet_name] = f"エラー: {str(e)}"
    
    return sheet_summaries


def create_final_summary(basename: str, sheet_summaries: Dict[str, str], 
                        client: AzureOpenAI) -> str:
    """
    全シート概要から最終的な解説を作成
    """
    summaries_text = []
    for sheet_name, summary in sheet_summaries.items():
        summaries_text.append(f"### {sheet_name}\n{summary}\n")
    
    combined_summaries = "\n".join(summaries_text)
    
    response = client.chat.completions.create(
        model=DEPLOYMENT_NAME,
        messages=[
            {"role": "system", "content": FINAL_SUMMARY_PROMPT.format(basename=basename)},
            {"role": "user", "content": combined_summaries}
        ],
        max_tokens=10000
    )
    return response.choices[0].message.content


def create_summary(rel_path: str, resource_dir: str, client: AzureOpenAI, 
                  use_chunking: bool = False, include_images: bool = True) -> str:
    """
    Excelファイルから解説を作成（2段階方式、画像対応）
    """
    basename = os.path.basename(rel_path)
    
    excel_path = os.path.join(resource_dir, f'{rel_path}.xlsx')
    if not os.path.exists(excel_path):
        excel_path = os.path.join(resource_dir, f'{rel_path}.xls')
    
    all_sheets = pd.read_excel(excel_path, sheet_name=None)
    print(f'  Found {len(all_sheets)} sheets')
    
    # 段階1: 各シートの概要を作成
    print(f'  Phase 1: Creating sheet summaries...')
    sheet_summaries = create_sheet_summaries(basename, all_sheets, excel_path, client, 
                                            use_chunking, include_images)
    
    # 段階2: 全体の解説を作成
    print(f'  Phase 2: Creating final summary...')
    time.sleep(WAIT_TIME_BETWEEN_SHEETS)
    final_summary = create_final_summary(basename, sheet_summaries, client)
    
    # 最終的なMarkdownを構築
    result_parts = [
        f"# {basename} - 機能解説\n\n",
        final_summary,
        "\n\n---\n\n",
        "## 各シート概要\n\n"
    ]
    
    for sheet_name, summary in sheet_summaries.items():
        result_parts.append(f"### {sheet_name}\n\n{summary}\n\n")
    
    return "".join(result_parts)


def main(use_chunking: bool = False, include_images: bool = True):
    """
    メイン処理
    """
    inputs = get_excel_files(INPUT_DIR)
    
    output_mds = []
    if os.path.exists(OUTPUT_DIR):
        for root, dirs, files in os.walk(OUTPUT_DIR):
            for file in files:
                if file.endswith('.md'):
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, OUTPUT_DIR)
                    rel_path_without_ext = os.path.splitext(rel_path)[0]
                    output_mds.append(rel_path_without_ext)
    
    undescribeds = [f for f in inputs if f not in output_mds]
    
    print(f'Found {len(inputs)} Excel files')
    print(f'Already processed: {len(output_mds)}')
    print(f'Processing {len(undescribeds)} files...')
    print(f'Chunking mode: {"ON" if use_chunking else "OFF"}')
    print(f'Include images: {"ON" if include_images else "OFF"}')
    
    if len(undescribeds) == 0:
        print('No files to process.')
        return
    
    client = get_client()
    
    for idx, undescribed in enumerate(undescribeds, 1):
        try:
            print(f'\n[{idx}/{len(undescribeds)}] 📄 Processing: {undescribed}')
            description = create_summary(undescribed, INPUT_DIR, client, use_chunking, include_images)
            
            output_path = os.path.join(OUTPUT_DIR, f'{undescribed}.md')
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(description)
            
            print(f'✓ Summary file created: {undescribed}')
            
            if idx < len(undescribeds):
                time.sleep(WAIT_TIME_BETWEEN_FILES)
            
        except Exception as e:
            print(f'✗ Error processing {undescribed}: {str(e)}')
    
    print(f'\n{"="*60}')
    print(f'Processing complete!')


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Create summary from Excel files with image support')
    parser.add_argument('--chunking', action='store_true', 
                       help='Enable chunking for large sheets')
    parser.add_argument('--no-images', action='store_true',
                       help='Disable image extraction')
    
    args = parser.parse_args()
    
    main(use_chunking=args.chunking, include_images=not args.no_images)
