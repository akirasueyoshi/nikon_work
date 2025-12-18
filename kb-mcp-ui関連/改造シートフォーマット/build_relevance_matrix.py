#!/usr/bin/env python3
"""
JSONファイルから関連度マトリクスを作成し、検索評価用の正解データを生成する
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
import csv


def load_document_graph(json_path):
    """JSONファイルからドキュメントグラフを読み込む"""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data


def build_adjacency_matrix(data):
    """
    隣接行列を構築
    
    Returns:
        matrix: numpy array (directed graph)
        docs: list of document IDs
    """
    # ドキュメントIDリストを作成
    docs = sorted([d['id'] for d in data['documents']])
    n = len(docs)
    doc_to_idx = {doc: i for i, doc in enumerate(docs)}
    
    # 隣接行列を初期化
    matrix = np.zeros((n, n))
    
    # リンク情報から行列を構築
    for link in data['links']:
        source = link['source']
        target = link['target']
        
        if source in doc_to_idx and target in doc_to_idx:
            i = doc_to_idx[source]
            j = doc_to_idx[target]
            matrix[i][j] = 1.0
    
    return matrix, docs


def calculate_relevance_matrix(adjacency, method="combined"):
    """
    複数の手法で関連度マトリクスを計算
    
    Parameters:
        adjacency: 隣接行列（directed graph）
        method: 計算方法
            - "direct": 直接リンクのみ
            - "bidirectional": 双方向リンク重視
            - "common_links": 共通リンク先（類似性）
            - "combined": 上記の組み合わせ（推奨）
    
    Returns:
        relevance_matrix: 関連度行列
    """
    n = adjacency.shape[0]
    
    if method == "direct":
        # 直接リンクのみ（そのまま）
        relevance = adjacency.copy()
        np.fill_diagonal(relevance, 1.0)
        
    elif method == "bidirectional":
        # 双方向性を考慮
        relevance = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                if i == j:
                    relevance[i][j] = 1.0
                else:
                    a_to_b = adjacency[i][j]
                    b_to_a = adjacency[j][i]
                    
                    if a_to_b and b_to_a:
                        relevance[i][j] = 1.0  # 相互リンク
                    elif a_to_b or b_to_a:
                        relevance[i][j] = 0.7  # 片方向リンク
        
    elif method == "common_links":
        # 共通リンク先による類似度（Jaccard係数）
        relevance = np.eye(n)
        for i in range(n):
            for j in range(i+1, n):
                links_i = set(np.where(adjacency[i] > 0)[0])
                links_j = set(np.where(adjacency[j] > 0)[0])
                
                union = links_i | links_j
                if len(union) > 0:
                    jaccard = len(links_i & links_j) / len(union)
                    relevance[i][j] = jaccard
                    relevance[j][i] = jaccard  # 対称
        
    elif method == "combined":
        # 複数の指標を組み合わせる
        # 1. 直接リンク
        m_direct = adjacency.copy()
        
        # 2. 双方向性
        m_bidirectional = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                if i == j:
                    m_bidirectional[i][j] = 1.0
                else:
                    a_to_b = adjacency[i][j]
                    b_to_a = adjacency[j][i]
                    
                    if a_to_b and b_to_a:
                        m_bidirectional[i][j] = 1.0
                    elif a_to_b or b_to_a:
                        m_bidirectional[i][j] = 0.5
        
        # 3. 共通リンク先
        m_common = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                if i == j:
                    m_common[i][j] = 1.0
                else:
                    links_i = set(np.where(adjacency[i] > 0)[0])
                    links_j = set(np.where(adjacency[j] > 0)[0])
                    
                    union = links_i | links_j
                    if len(union) > 0:
                        m_common[i][j] = len(links_i & links_j) / len(union)
        
        # 重み付け統合（調整可能）
        relevance = (m_direct * 0.5 +           # 直接リンクを最重視
                     m_bidirectional * 0.3 +     # 双方向性
                     m_common * 0.2)             # 共通性
        
        # 正規化（0-1に収める）
        relevance = np.clip(relevance, 0, 1)
        
        # 対角を1にする
        np.fill_diagonal(relevance, 1.0)
    
    else:
        raise ValueError(f"Unknown method: {method}")
    
    return relevance


def create_ground_truth(relevance_matrix, docs, threshold=0.3, top_k=10):
    """
    検索評価用の正解データを生成
    
    Parameters:
        relevance_matrix: 関連度行列
        docs: ドキュメントIDリスト
        threshold: 関連ありとみなす閾値
        top_k: 各クエリについて保存する上位K件
    
    Returns:
        ground_truth: list of dicts
    """
    ground_truth = []
    
    for i, query_doc in enumerate(docs):
        # 関連度でソート（自分自身を除く）
        relevance_scores = [
            (docs[j], relevance_matrix[i][j])
            for j in range(len(docs)) if j != i
        ]
        relevance_scores.sort(key=lambda x: x[1], reverse=True)
        
        # 閾値以上のものを抽出
        relevant_docs = [
            {"doc_id": doc, "relevance": round(float(score), 3)}
            for doc, score in relevance_scores
            if score >= threshold
        ]
        
        # Top-Kに制限
        relevant_docs_topk = relevant_docs[:top_k]
        
        ground_truth.append({
            "query_doc": query_doc,
            "relevant_docs": relevant_docs_topk,
            "total_relevant": len(relevant_docs),
            "threshold": threshold
        })
    
    return ground_truth


def save_results(relevance_matrix, docs, ground_truth, output_dir="relevance_results"):
    """
    関連度マトリクスと正解データを保存
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. 関連度マトリクス（CSV）
    df_matrix = pd.DataFrame(relevance_matrix, index=docs, columns=docs)
    matrix_path = output_dir / f"relevance_matrix_{timestamp}.csv"
    df_matrix.to_csv(matrix_path, encoding="utf-8-sig")
    print(f"✓ Saved relevance matrix: {matrix_path}")
    
    # 2. エッジリスト（閾値以上のみ）
    edges = []
    threshold = 0.3
    for i, doc_a in enumerate(docs):
        for j, doc_b in enumerate(docs):
            if i != j and relevance_matrix[i][j] >= threshold:
                edges.append({
                    "source": doc_a,
                    "target": doc_b,
                    "relevance": round(float(relevance_matrix[i][j]), 3)
                })
    
    edge_path = output_dir / f"relevance_edges_{timestamp}.csv"
    with open(edge_path, "w", encoding="utf-8-sig", newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['source', 'target', 'relevance'])
        writer.writeheader()
        writer.writerows(edges)
    print(f"✓ Saved edge list: {edge_path}")
    
    # 3. 正解データ（JSON）
    gt_path = output_dir / f"ground_truth_{timestamp}.json"
    with open(gt_path, "w", encoding="utf-8") as f:
        json.dump(ground_truth, f, ensure_ascii=False, indent=2)
    print(f"✓ Saved ground truth: {gt_path}")
    
    # 4. 統計サマリー
    upper_triangle = relevance_matrix[np.triu_indices_from(relevance_matrix, k=1)]
    
    summary = {
        "timestamp": timestamp,
        "total_documents": len(docs),
        "total_edges_above_threshold": len(edges),
        "threshold": threshold,
        "statistics": {
            "mean_relevance": float(np.mean(upper_triangle)),
            "median_relevance": float(np.median(upper_triangle)),
            "std_relevance": float(np.std(upper_triangle)),
            "max_relevance": float(np.max(upper_triangle)),
            "min_relevance": float(np.min(upper_triangle))
        },
        "ground_truth_stats": {
            "total_queries": len(ground_truth),
            "avg_relevant_docs_per_query": np.mean([g['total_relevant'] for g in ground_truth]),
            "queries_with_no_relevant": len([g for g in ground_truth if g['total_relevant'] == 0])
        }
    }
    
    summary_path = output_dir / f"summary_{timestamp}.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"✓ Saved summary: {summary_path}")
    
    # コンソールにサマリー表示
    print("\n" + "="*60)
    print("RELEVANCE MATRIX SUMMARY")
    print("="*60)
    print(f"Documents:                 {summary['total_documents']}")
    print(f"Edges (≥{threshold}):           {summary['total_edges_above_threshold']}")
    print(f"Mean relevance:            {summary['statistics']['mean_relevance']:.3f}")
    print(f"Median relevance:          {summary['statistics']['median_relevance']:.3f}")
    print(f"Std relevance:             {summary['statistics']['std_relevance']:.3f}")
    print(f"\nGround Truth Stats:")
    print(f"Total queries:             {summary['ground_truth_stats']['total_queries']}")
    print(f"Avg relevant docs/query:   {summary['ground_truth_stats']['avg_relevant_docs_per_query']:.1f}")
    print(f"Queries with no relevant:  {summary['ground_truth_stats']['queries_with_no_relevant']}")
    print("="*60)
    
    return matrix_path, gt_path, timestamp


def create_excel_report(document_graph_json, relevance_matrix_csv, ground_truth_json, output_path):
    """
    関連度マトリクスと詳細情報をエクセルファイルに出力
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Warning: openpyxl not available, skipping Excel report")
        return None
    
    print(f"\nCreating Excel report...")
    
    # データ読み込み
    with open(document_graph_json, 'r', encoding='utf-8') as f:
        doc_graph = json.load(f)
    
    df_matrix = pd.read_csv(relevance_matrix_csv, index_col=0, encoding='utf-8-sig')
    
    with open(ground_truth_json, 'r', encoding='utf-8') as f:
        ground_truth = json.load(f)
    
    # Workbook作成
    wb = Workbook()
    wb.remove(wb.active)
    
    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: サマリー
    ws_summary = wb.create_sheet("サマリー", 0)
    ws_summary.cell(1, 1, "関連度マトリクス分析サマリー").font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:B1')
    
    metadata = doc_graph['metadata']
    summary_data = [
        ["作成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["抽出日時", metadata.get('extraction_date', 'N/A')],
        ["", ""],
        ["総資料数", metadata['total_documents']],
        ["マッチしたリンク数", metadata['total_matched_links']],
        ["未マッチリンク数", metadata['total_unmatched_links']],
        ["", ""],
        ["関連度 ≥ 0.7 のペア数", len([(i,j) for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j and df_matrix.iloc[i,j] >= 0.7])],
        ["関連度 ≥ 0.3 のペア数", len([(i,j) for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j and df_matrix.iloc[i,j] >= 0.3])],
    ]
    
    for idx, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(idx, 1, label).font = Font(bold=True)
        ws_summary.cell(idx, 2, value)
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # Sheet 2: 関連度マトリクス
    ws_matrix = wb.create_sheet("関連度マトリクス")
    ws_matrix.cell(1, 1, "資料名")
    for col_idx, col_name in enumerate(df_matrix.columns, 2):
        cell = ws_matrix.cell(1, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, (index_name, row_data) in enumerate(df_matrix.iterrows(), 2):
        cell = ws_matrix.cell(row_idx, 1, index_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for col_idx, value in enumerate(row_data, 2):
            cell = ws_matrix.cell(row_idx, col_idx, round(float(value), 3))
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_idx - 1 == col_idx - 1:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif value >= 0.7:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                cell.font = Font(bold=True)
            elif value >= 0.3:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws_matrix.column_dimensions['A'].width = 40
    for col_idx in range(2, len(df_matrix.columns) + 2):
        ws_matrix.column_dimensions[chr(64 + col_idx)].width = 15
    
    # Sheet 3: ドキュメントとリンク
    ws_links = wb.create_sheet("ドキュメントとリンク")
    headers = ["資料名", "ファイル名", "抽出リンク数", "マッチしたリンク数", "リンク先資料"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_links.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    for row_idx, doc in enumerate(doc_graph['documents'], 2):
        matched_links = [link['target'] for link in doc_graph['links'] if link['source'] == doc['id']]
        ws_links.cell(row_idx, 1, doc['id']).border = border
        ws_links.cell(row_idx, 2, doc['filename']).border = border
        ws_links.cell(row_idx, 3, doc['extracted_links_count']).border = border
        ws_links.cell(row_idx, 4, len(matched_links)).border = border
        ws_links.cell(row_idx, 5, ", ".join(matched_links) if matched_links else "(なし)").border = border
    
    ws_links.column_dimensions['A'].width = 40
    ws_links.column_dimensions['B'].width = 40
    ws_links.column_dimensions['C'].width = 15
    ws_links.column_dimensions['D'].width = 18
    ws_links.column_dimensions['E'].width = 60
    
    # Sheet 4: リンク詳細
    ws_edges = wb.create_sheet("リンク詳細")
    headers = ["リンク元", "リンク先", "元テキスト", "マッチタイプ", "関連度"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_edges.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, link in enumerate(doc_graph['links'], 2):
        try:
            relevance = df_matrix.loc[link['source'], link['target']]
        except:
            relevance = 0.0
        
        ws_edges.cell(row_idx, 1, link['source']).border = border
        ws_edges.cell(row_idx, 2, link['target']).border = border
        ws_edges.cell(row_idx, 3, link['original_text']).border = border
        ws_edges.cell(row_idx, 4, link['match_type']).border = border
        cell = ws_edges.cell(row_idx, 5, round(float(relevance), 3))
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        if relevance >= 0.7:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif relevance >= 0.3:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws_edges.column_dimensions['A'].width = 40
    ws_edges.column_dimensions['B'].width = 40
    ws_edges.column_dimensions['C'].width = 40
    ws_edges.column_dimensions['D'].width = 15
    ws_edges.column_dimensions['E'].width = 12
    
    # Sheet 5: 未マッチリンク
    if doc_graph.get('unmatched_links'):
        ws_unmatched = wb.create_sheet("未マッチリンク")
        headers = ["リンク元", "元テキスト", "正規化後"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_unmatched.cell(1, col_idx, header)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, link in enumerate(doc_graph['unmatched_links'], 2):
            ws_unmatched.cell(row_idx, 1, link['source']).border = border
            ws_unmatched.cell(row_idx, 2, link['original_text']).border = border
            ws_unmatched.cell(row_idx, 3, link['normalized']).border = border
        
        ws_unmatched.column_dimensions['A'].width = 40
        ws_unmatched.column_dimensions['B'].width = 50
        ws_unmatched.column_dimensions['C'].width = 50
    
    # Sheet 6: 正解データ
    ws_gt = wb.create_sheet("正解データサマリー")
    headers = ["クエリ資料", "関連資料数", "上位関連資料", "関連度", "閾値"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_gt.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    row_idx = 2
    for gt_entry in ground_truth:
        if gt_entry['relevant_docs']:
            for i, rel_doc in enumerate(gt_entry['relevant_docs']):
                if i == 0:
                    ws_gt.cell(row_idx, 1, gt_entry['query_doc']).border = border
                    ws_gt.cell(row_idx, 2, gt_entry['total_relevant']).border = border
                else:
                    ws_gt.cell(row_idx, 1, "").border = border
                    ws_gt.cell(row_idx, 2, "").border = border
                
                ws_gt.cell(row_idx, 3, rel_doc['doc_id']).border = border
                cell = ws_gt.cell(row_idx, 4, rel_doc['relevance'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                if i == 0:
                    ws_gt.cell(row_idx, 5, gt_entry['threshold']).border = border
                else:
                    ws_gt.cell(row_idx, 5, "").border = border
                row_idx += 1
        else:
            ws_gt.cell(row_idx, 1, gt_entry['query_doc']).border = border
            ws_gt.cell(row_idx, 2, 0).border = border
            ws_gt.cell(row_idx, 3, "(関連資料なし)").border = border
            ws_gt.cell(row_idx, 4, "").border = border
            ws_gt.cell(row_idx, 5, gt_entry['threshold']).border = border
            row_idx += 1
    
    ws_gt.column_dimensions['A'].width = 40
    ws_gt.column_dimensions['B'].width = 15
    ws_gt.column_dimensions['C'].width = 40
    ws_gt.column_dimensions['D'].width = 12
    ws_gt.column_dimensions['E'].width = 12
    
    # 保存
    wb.save(output_path)
    print(f"✓ Saved Excel report: {output_path}")
    return output_path


def visualize_matrix(relevance_matrix, docs, output_path):
    """
    関連度マトリクスをヒートマップで可視化
    （30x30以下の場合のみ）
    """
    if len(docs) > 30:
        print(f"Skipping heatmap (too many documents: {len(docs)})")
        return
    
    try:
        import matplotlib
        matplotlib.use('Agg')  # GUIなし環境用
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        plt.figure(figsize=(16, 14))
        sns.heatmap(relevance_matrix, 
                   xticklabels=docs,
                   yticklabels=docs,
                   cmap="YlOrRd",
                   vmin=0, vmax=1,
                   square=True,
                   cbar_kws={'label': 'Relevance'})
        plt.title("Document Relevance Matrix")
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        print(f"✓ Saved heatmap: {output_path}")
    except ImportError:
        print("Note: matplotlib/seaborn not available, skipping visualization")


def main():
    """メイン実行関数"""
    import sys
    
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    else:
        json_path = input("Enter path to document_graph JSON file: ").strip()
    
    json_path = Path(json_path)
    if not json_path.exists():
        print(f"Error: File not found: {json_path}")
        return
    
    print("\n" + "="*60)
    print("RELEVANCE MATRIX BUILDER")
    print("="*60)
    
    # JSONを読み込み
    print(f"\nLoading: {json_path}")
    data = load_document_graph(json_path)
    
    # 隣接行列を構築
    print("Building adjacency matrix...")
    adjacency, docs = build_adjacency_matrix(data)
    print(f"  Documents: {len(docs)}")
    print(f"  Links: {int(np.sum(adjacency))}")
    
    # 関連度マトリクスを計算
    print("\nCalculating relevance matrix (method: combined)...")
    relevance_matrix = calculate_relevance_matrix(adjacency, method="combined")
    
    # 正解データを生成
    print("Creating ground truth data...")
    ground_truth = create_ground_truth(relevance_matrix, docs, threshold=0.3, top_k=10)
    
    # 結果を保存
    print("\nSaving results...")
    matrix_path, gt_path, timestamp = save_results(relevance_matrix, docs, ground_truth)
    
    # エクセルレポート作成
    output_dir = matrix_path.parent
    excel_path = output_dir / f"relevance_report_{timestamp}.xlsx"
    try:
        create_excel_report(json_path, matrix_path, gt_path, excel_path)
    except Exception as e:
        print(f"Warning: Could not create Excel report: {e}")
    
    # 可視化
    heatmap_path = output_dir / f"heatmap_{timestamp}.png"
    visualize_matrix(relevance_matrix, docs, heatmap_path)
    
    print(f"\n✓ Processing completed successfully!")
    print(f"\nMain outputs:")
    print(f"  - Relevance matrix CSV: {matrix_path}")
    print(f"  - Ground truth JSON:    {gt_path}")
    if excel_path.exists():
        print(f"  - Excel report:         {excel_path}")
    

if __name__ == "__main__":
    main()