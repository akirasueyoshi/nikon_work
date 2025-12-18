#!/usr/bin/env python3
"""
JSONファイルとCSVファイルから関連度マトリクスをエクセル形式で出力
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import sys


def create_excel_report(document_graph_json, relevance_matrix_csv, ground_truth_json, output_path):
    """
    関連度マトリクスと詳細情報をエクセルファイルに出力
    
    Parameters:
        document_graph_json: ドキュメントグラフのJSONファイルパス
        relevance_matrix_csv: 関連度マトリクスのCSVファイルパス
        ground_truth_json: 正解データのJSONファイルパス
        output_path: 出力先エクセルファイルパス
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    print(f"Loading data files...")
    
    # データ読み込み
    with open(document_graph_json, 'r', encoding='utf-8') as f:
        doc_graph = json.load(f)
    
    df_matrix = pd.read_csv(relevance_matrix_csv, index_col=0, encoding='utf-8-sig')
    
    with open(ground_truth_json, 'r', encoding='utf-8') as f:
        ground_truth = json.load(f)
    
    # Workbook作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    
    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== Sheet 1: 関連度マトリクス ====================
    print("Creating Sheet 1: Relevance Matrix...")
    ws_matrix = wb.create_sheet("関連度マトリクス")
    
    # ヘッダー行を追加
    ws_matrix.cell(1, 1, "資料名")
    for col_idx, col_name in enumerate(df_matrix.columns, 2):
        cell = ws_matrix.cell(1, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # データ行を追加
    for row_idx, (index_name, row_data) in enumerate(df_matrix.iterrows(), 2):
        # 行ヘッダー
        cell = ws_matrix.cell(row_idx, 1, index_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # データセル
        for col_idx, value in enumerate(row_data, 2):
            cell = ws_matrix.cell(row_idx, col_idx, round(float(value), 3))
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 対角線（自分自身）
            if row_idx - 1 == col_idx - 1:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            # 関連度が高い（0.7以上）
            elif value >= 0.7:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                cell.font = Font(bold=True)
            # 関連度が中程度（0.3-0.7）
            elif value >= 0.3:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 列幅調整
    ws_matrix.column_dimensions['A'].width = 40
    for col_idx in range(2, len(df_matrix.columns) + 2):
        ws_matrix.column_dimensions[chr(64 + col_idx)].width = 15
    
    # ==================== Sheet 2: ドキュメントとリンク一覧 ====================
    print("Creating Sheet 2: Documents and Links...")
    ws_links = wb.create_sheet("ドキュメントとリンク")
    
    # ヘッダー
    headers = ["資料名", "ファイル名", "抽出リンク数", "マッチしたリンク数", "リンク先資料"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_links.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # ドキュメントごとのリンク情報
    row_idx = 2
    for doc in doc_graph['documents']:
        doc_id = doc['id']
        
        # リンク先を取得
        matched_links = [link['target'] for link in doc_graph['links'] if link['source'] == doc_id]
        
        ws_links.cell(row_idx, 1, doc_id).border = border
        ws_links.cell(row_idx, 2, doc['filename']).border = border
        ws_links.cell(row_idx, 3, doc['extracted_links_count']).border = border
        ws_links.cell(row_idx, 4, len(matched_links)).border = border
        ws_links.cell(row_idx, 5, ", ".join(matched_links) if matched_links else "(なし)").border = border
        
        row_idx += 1
    
    # 列幅調整
    ws_links.column_dimensions['A'].width = 40
    ws_links.column_dimensions['B'].width = 40
    ws_links.column_dimensions['C'].width = 15
    ws_links.column_dimensions['D'].width = 18
    ws_links.column_dimensions['E'].width = 60
    
    # ==================== Sheet 3: リンク詳細（エッジリスト） ====================
    print("Creating Sheet 3: Link Details...")
    ws_edges = wb.create_sheet("リンク詳細")
    
    # ヘッダー
    headers = ["リンク元", "リンク先", "元テキスト", "マッチタイプ", "関連度"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_edges.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # リンクデータ
    row_idx = 2
    for link in doc_graph['links']:
        source = link['source']
        target = link['target']
        
        # 関連度を取得
        try:
            relevance = df_matrix.loc[source, target]
        except:
            relevance = 0.0
        
        ws_edges.cell(row_idx, 1, source).border = border
        ws_edges.cell(row_idx, 2, target).border = border
        ws_edges.cell(row_idx, 3, link['original_text']).border = border
        ws_edges.cell(row_idx, 4, link['match_type']).border = border
        cell = ws_edges.cell(row_idx, 5, round(float(relevance), 3))
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # 関連度による色付け
        if relevance >= 0.7:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif relevance >= 0.3:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        row_idx += 1
    
    # 列幅調整
    ws_edges.column_dimensions['A'].width = 40
    ws_edges.column_dimensions['B'].width = 40
    ws_edges.column_dimensions['C'].width = 40
    ws_edges.column_dimensions['D'].width = 15
    ws_edges.column_dimensions['E'].width = 12
    
    # ==================== Sheet 4: 未マッチリンク ====================
    print("Creating Sheet 4: Unmatched Links...")
    ws_unmatched = wb.create_sheet("未マッチリンク")
    
    # ヘッダー
    headers = ["リンク元", "元テキスト", "正規化後"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_unmatched.cell(1, col_idx, header)
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 未マッチデータ
    row_idx = 2
    for link in doc_graph.get('unmatched_links', []):
        ws_unmatched.cell(row_idx, 1, link['source']).border = border
        ws_unmatched.cell(row_idx, 2, link['original_text']).border = border
        ws_unmatched.cell(row_idx, 3, link['normalized']).border = border
        row_idx += 1
    
    # 列幅調整
    ws_unmatched.column_dimensions['A'].width = 40
    ws_unmatched.column_dimensions['B'].width = 50
    ws_unmatched.column_dimensions['C'].width = 50
    
    # ==================== Sheet 5: 正解データサマリー ====================
    print("Creating Sheet 5: Ground Truth Summary...")
    ws_gt = wb.create_sheet("正解データサマリー")
    
    # ヘッダー
    headers = ["クエリ資料", "関連資料数", "上位関連資料", "関連度", "閾値"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_gt.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # 正解データ
    row_idx = 2
    for gt_entry in ground_truth:
        query_doc = gt_entry['query_doc']
        total_relevant = gt_entry['total_relevant']
        threshold = gt_entry['threshold']
        
        if gt_entry['relevant_docs']:
            # 各関連資料を1行ずつ
            for i, rel_doc in enumerate(gt_entry['relevant_docs']):
                if i == 0:
                    ws_gt.cell(row_idx, 1, query_doc).border = border
                    ws_gt.cell(row_idx, 2, total_relevant).border = border
                else:
                    ws_gt.cell(row_idx, 1, "").border = border
                    ws_gt.cell(row_idx, 2, "").border = border
                
                ws_gt.cell(row_idx, 3, rel_doc['doc_id']).border = border
                cell = ws_gt.cell(row_idx, 4, rel_doc['relevance'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                if i == 0:
                    ws_gt.cell(row_idx, 5, threshold).border = border
                else:
                    ws_gt.cell(row_idx, 5, "").border = border
                
                row_idx += 1
        else:
            # 関連資料なし
            ws_gt.cell(row_idx, 1, query_doc).border = border
            ws_gt.cell(row_idx, 2, 0).border = border
            ws_gt.cell(row_idx, 3, "(関連資料なし)").border = border
            ws_gt.cell(row_idx, 4, "").border = border
            ws_gt.cell(row_idx, 5, threshold).border = border
            row_idx += 1
    
    # 列幅調整
    ws_gt.column_dimensions['A'].width = 40
    ws_gt.column_dimensions['B'].width = 15
    ws_gt.column_dimensions['C'].width = 40
    ws_gt.column_dimensions['D'].width = 12
    ws_gt.column_dimensions['E'].width = 12
    
    # ==================== Sheet 6: サマリー ====================
    print("Creating Sheet 6: Summary...")
    ws_summary = wb.create_sheet("サマリー", 0)  # 最初のシートに
    
    # タイトル
    ws_summary.cell(1, 1, "関連度マトリクス分析サマリー")
    ws_summary.cell(1, 1).font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:B1')
    
    # メタデータ
    metadata = doc_graph['metadata']
    summary_data = [
        ["作成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["抽出日時", metadata.get('extraction_date', 'N/A')],
        ["", ""],
        ["総資料数", metadata['total_documents']],
        ["マッチしたリンク数", metadata['total_matched_links']],
        ["未マッチリンク数", metadata['total_unmatched_links']],
        ["", ""],
        ["関連度 ≥ 0.7 のペア数", len([(i,j) for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j and df_matrix.iloc[i,j] >= 0.7])],
        ["関連度 ≥ 0.3 のペア数", len([(i,j) for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j and df_matrix.iloc[i,j] >= 0.3])],
        ["平均関連度", round(sum([df_matrix.iloc[i,j] for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j]) / max(1, len([(i,j) for i in range(len(df_matrix)) for j in range(len(df_matrix)) if i<j])), 3)],
    ]
    
    row_idx = 3
    for label, value in summary_data:
        ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
        ws_summary.cell(row_idx, 2, value)
        row_idx += 1
    
    # 列幅調整
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # 保存
    print(f"Saving Excel file: {output_path}")
    wb.save(output_path)
    print(f"✓ Excel report created successfully!")
    
    return output_path


def main():
    """メイン実行関数"""
    if len(sys.argv) < 4:
        print("Usage: python3 create_excel_report.py <document_graph.json> <relevance_matrix.csv> <ground_truth.json> [output.xlsx]")
        print("\nExample:")
        print("  python3 create_excel_report.py \\")
        print("    extraction_results/document_graph_*.json \\")
        print("    relevance_results/relevance_matrix_*.csv \\")
        print("    relevance_results/ground_truth_*.json")
        return
    
    document_graph_json = Path(sys.argv[1])
    relevance_matrix_csv = Path(sys.argv[2])
    ground_truth_json = Path(sys.argv[3])
    
    # 出力パス
    if len(sys.argv) > 4:
        output_path = Path(sys.argv[4])
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(f"relevance_report_{timestamp}.xlsx")
    
    # ファイル存在チェック
    for filepath in [document_graph_json, relevance_matrix_csv, ground_truth_json]:
        if not filepath.exists():
            print(f"Error: File not found: {filepath}")
            return
    
    print("\n" + "="*60)
    print("EXCEL REPORT GENERATOR")
    print("="*60)
    print(f"Input files:")
    print(f"  - Document graph: {document_graph_json.name}")
    print(f"  - Relevance matrix: {relevance_matrix_csv.name}")
    print(f"  - Ground truth: {ground_truth_json.name}")
    print(f"Output file: {output_path}")
    print("="*60 + "\n")
    
    # エクセルレポート作成
    result_path = create_excel_report(
        document_graph_json,
        relevance_matrix_csv,
        ground_truth_json,
        output_path
    )
    
    print("\n" + "="*60)
    print("✓ Report generation completed!")
    print(f"Output: {result_path}")
    print("="*60)
    

if __name__ == "__main__":
    main()
